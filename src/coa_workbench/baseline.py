from __future__ import annotations

import csv
import hashlib
import json
import re
import warnings
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

FORMULA_FUNCTION = re.compile(r"\b([A-Z][A-Z0-9.]*)\s*\(")
EXCEL_ERRORS = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A", "#GETTING_DATA"}
PROJECT_DOCUMENT_TIMELINE_EVENTS = 12_147_472


@dataclass(frozen=True)
class FrozenFile:
    path: str
    size_bytes: int
    sha256: str
    modified_at: str


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, tuple):
        return list(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def freeze_file(path: Path, *, relative_to: Path | None = None) -> FrozenFile:
    stat = path.stat()
    display_path = path.relative_to(relative_to) if relative_to and path.is_relative_to(relative_to) else path
    return FrozenFile(
        path=str(display_path),
        size_bytes=stat.st_size,
        sha256=sha256_file(path),
        modified_at=datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(),
    )


def _load_pair(workbook_path: Path):
    captured: list[str] = []
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        formula_wb = load_workbook(workbook_path, data_only=False, keep_links=True)
        value_wb = load_workbook(workbook_path, data_only=True, keep_links=True)
        captured.extend(str(item.message) for item in caught)
    return formula_wb, value_wb, sorted(set(captured))


def _package_extension_inventory(workbook_path: Path) -> dict[str, Any]:
    markers = Counter()
    interesting_parts: list[str] = []
    with zipfile.ZipFile(workbook_path) as archive:
        part_names = archive.namelist()
        for name in part_names:
            lower = name.lower()
            if any(token in lower for token in ("externallink", "connections", "querytable", "pivot", "vba", "calcchain")):
                interesting_parts.append(name)
            if not name.endswith(".xml"):
                continue
            text = archive.read(name).decode("utf-8", errors="ignore")
            for marker in (
                "<extLst",
                "x14:dataValidations",
                "x14:conditionalFormatting",
                "x15:",
                "xr:",
            ):
                count = text.count(marker)
                if count:
                    markers[marker] += count
    return {
        "part_count": len(part_names),
        "interesting_parts": sorted(interesting_parts),
        "extension_markers": dict(markers),
    }


def build_inventory(workbook_path: Path) -> tuple[dict[str, Any], Any, Any]:
    formula_wb, value_wb, load_warnings = _load_pair(workbook_path)
    package = _package_extension_inventory(workbook_path)
    sheets: list[dict[str, Any]] = []
    total_formulas = 0
    total_cached_errors = 0
    all_error_cells: list[dict[str, Any]] = []

    for formula_ws in formula_wb.worksheets:
        value_ws = value_wb[formula_ws.title]
        nonempty = 0
        formula_count = 0
        functions: Counter[str] = Counter()
        cached_errors: Counter[str] = Counter()
        error_cells: list[dict[str, Any]] = []
        formula_samples: list[dict[str, Any]] = []

        for row in formula_ws.iter_rows():
            for formula_cell in row:
                formula_value = formula_cell.value
                cached_value = value_ws[formula_cell.coordinate].value
                if formula_value is None:
                    continue
                nonempty += 1
                if isinstance(formula_value, str) and formula_value.startswith("="):
                    formula_count += 1
                    functions.update(FORMULA_FUNCTION.findall(formula_value.upper()))
                    if len(formula_samples) < 5:
                        formula_samples.append(
                            {
                                "cell": formula_cell.coordinate,
                                "formula": formula_value,
                                "cached_value": _json_value(cached_value),
                            }
                        )
                    if isinstance(cached_value, str) and cached_value in EXCEL_ERRORS:
                        cached_errors[cached_value] += 1
                        item = {
                            "sheet": formula_ws.title,
                            "cell": formula_cell.coordinate,
                            "error": cached_value,
                            "formula": formula_value,
                        }
                        error_cells.append(item)
                        all_error_cells.append(item)

        total_formulas += formula_count
        total_cached_errors += sum(cached_errors.values())
        tables = []
        for table_name in formula_ws.tables.keys():
            table = formula_ws.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            tables.append(
                {
                    "name": table_name,
                    "ref": table.ref,
                    "data_rows": max_row - min_row,
                    "columns": max_col - min_col + 1,
                }
            )
        sheets.append(
            {
                "name": formula_ws.title,
                "state": formula_ws.sheet_state,
                "max_row": formula_ws.max_row,
                "max_column": formula_ws.max_column,
                "nonempty_cells": nonempty,
                "formula_count": formula_count,
                "cached_formula_errors": dict(cached_errors),
                "cached_error_cells": error_cells,
                "top_formula_functions": functions.most_common(15),
                "formula_samples": formula_samples,
                "tables": tables,
                "merged_ranges": len(formula_ws.merged_cells.ranges),
                "data_validations_loaded": len(formula_ws.data_validations.dataValidation)
                if formula_ws.data_validations
                else 0,
                "conditional_formatting_rules_loaded": len(formula_ws.conditional_formatting),
                "charts": len(formula_ws._charts),
                "images": len(formula_ws._images),
                "freeze_panes": str(formula_ws.freeze_panes) if formula_ws.freeze_panes else None,
                "auto_filter": formula_ws.auto_filter.ref,
            }
        )

    calc = formula_wb.calculation
    inventory = {
        "schema_version": 1,
        "generated_at": datetime.now().astimezone().isoformat(),
        "source": asdict(freeze_file(workbook_path)),
        "workbook": {
            "sheet_count": len(formula_wb.sheetnames),
            "sheet_names": formula_wb.sheetnames,
            "defined_name_count": len(formula_wb.defined_names),
            "defined_names": [
                {
                    "name": name,
                    "type": item.type,
                    "attr_text": item.attr_text,
                    "local_sheet_id": item.localSheetId,
                    "hidden": item.hidden,
                }
                for name, item in formula_wb.defined_names.items()
            ],
            "properties": {
                "creator": formula_wb.properties.creator,
                "last_modified_by": formula_wb.properties.lastModifiedBy,
                "created": _json_value(formula_wb.properties.created),
                "modified": _json_value(formula_wb.properties.modified),
            },
            "calculation": {
                "calc_id": calc.calcId,
                "calc_mode": calc.calcMode,
                "full_calc_on_load": calc.fullCalcOnLoad,
                "force_full_calc": calc.forceFullCalc,
            },
            "total_formula_count": total_formulas,
            "total_cached_formula_errors": total_cached_errors,
            "cached_formula_error_cells": all_error_cells,
            "load_warnings": load_warnings,
            "openpyxl_resave_safe": not load_warnings and not package["extension_markers"],
            "package": package,
        },
        "sheets": sheets,
    }
    return inventory, formula_wb, value_wb


def _safe_filename(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("_")
    return value or "table"


def export_tables(formula_wb: Any, value_wb: Any, output_dir: Path) -> list[dict[str, Any]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    exports: list[dict[str, Any]] = []
    for formula_ws in formula_wb.worksheets:
        value_ws = value_wb[formula_ws.title]
        for table_name in formula_ws.tables.keys():
            table = formula_ws.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            target = output_dir / f"{_safe_filename(table_name)}.csv"
            with target.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle, lineterminator="\n")
                for row_no in range(min_row, max_row + 1):
                    writer.writerow(
                        [
                            _json_value(value_ws.cell(row_no, col_no).value)
                            for col_no in range(min_col, max_col + 1)
                        ]
                    )
            exports.append(
                {
                    "sheet": formula_ws.title,
                    "table": table_name,
                    "ref": table.ref,
                    "data_rows": max_row - min_row,
                    "columns": max_col - min_col + 1,
                    "path": str(target.relative_to(output_dir.parent)),
                    "sha256": sha256_file(target),
                }
            )
    return exports


def _cell(value_wb: Any, sheet: str, coordinate: str) -> Any:
    return _json_value(value_wb[sheet][coordinate].value)


def build_default_fixture(workbook_sha256: str, value_wb: Any, inventory: dict[str, Any]) -> dict[str, Any]:
    constructor = value_wb["РЕЙД-КОНСТРУКТОР"]
    comparison = value_wb["СРАВНЕНИЕ СПЕКОВ"]
    slots = []
    for row_no in range(6, 31):
        slots.append(
            {
                "slot_no": constructor[f"A{row_no}"].value,
                "player": constructor[f"B{row_no}"].value,
                "class": constructor[f"C{row_no}"].value,
                "spec": constructor[f"D{row_no}"].value,
                "role": constructor[f"E{row_no}"].value,
                "quantity": constructor[f"F{row_no}"].value,
                "validation": constructor[f"G{row_no}"].value,
                "counted": constructor[f"H{row_no}"].value,
            }
        )
    effects = []
    for row_no in range(6, 51):
        effects.append(
            {
                "category": constructor[f"I{row_no}"].value,
                "effect": constructor[f"J{row_no}"].value,
                "in_raid": constructor[f"K{row_no}"].value,
                "status": constructor[f"L{row_no}"].value,
                "priority": constructor[f"M{row_no}"].value,
                "weight": constructor[f"N{row_no}"].value,
            }
        )
    top_candidates = []
    for row_no in range(35, 45):
        top_candidates.append(
            {
                "rank": constructor[f"A{row_no}"].value,
                "class": constructor[f"B{row_no}"].value,
                "spec": constructor[f"C{row_no}"].value,
                "role": constructor[f"D{row_no}"].value,
                "critical": constructor[f"E{row_no}"].value,
                "important": constructor[f"F{row_no}"].value,
                "logs": constructor[f"G{row_no}"].value,
                "total": constructor[f"H{row_no}"].value,
            }
        )
    return {
        "schema_version": 1,
        "fixture_id": "workbook_v9_observed_default_state",
        "fixture_status": "observed_not_approved_25_roster",
        "source_workbook_sha256": workbook_sha256,
        "source_sheet": "РЕЙД-КОНСТРУКТОР",
        "context": {
            "recommendation_role": constructor["C3"].value,
            "scoring_profile": constructor["P2"].value,
            "uniqueness": constructor["R2"].value,
            "max_spec": constructor["T2"].value,
            "max_class": constructor["V2"].value,
            "raid_profile": constructor["P3"].value,
            "target_size": constructor["R3"].value,
        },
        "summary": {
            "counted_players": constructor["E3"].value,
            "missing_critical": constructor["G3"].value,
            "weighted_coverage": constructor["H3"].value,
            "missing_effects": constructor["I3"].value,
            "missing_critical_duplicate": constructor["K3"].value,
            "readiness": constructor["M3"].value,
            "remaining_slots": constructor["T3"].value,
            "top_recommendation": constructor["O6"].value,
            "top_recommendation_reason": constructor["O8"].value,
        },
        "slots": slots,
        "effects": effects,
        "top_candidates": top_candidates,
        "log_summary": {
            "canonical_kills_20_30": _cell(value_wb, "ЛОГИ_Сводка", "E5"),
            "kills_exactly_25": _cell(value_wb, "ЛОГИ_Сводка", "E6"),
            "workbook_timeline_events": _cell(value_wb, "ЛОГИ_Сводка", "E7"),
            "project_document_alternate_timeline_events": PROJECT_DOCUMENT_TIMELINE_EVENTS,
            "difference": _cell(value_wb, "ЛОГИ_Сводка", "E7") - PROJECT_DOCUMENT_TIMELINE_EVENTS,
            "aura_families": _cell(value_wb, "ЛОГИ_Сводка", "E8"),
            "effect_provider_links": _cell(value_wb, "ЛОГИ_Сводка", "E9"),
            "possible_exclusive_pairs": _cell(value_wb, "ЛОГИ_Сводка", "E10"),
            "status": "declared_not_canonical_until_raw_recount",
        },
        "comparison_default": {
            "variant_a": {
                "class": comparison["B5"].value,
                "spec": comparison["B6"].value,
                "role": comparison["B7"].value,
                "new_effects": comparison["B8"].value,
                "priority_score": comparison["B9"].value,
            },
            "variant_b": {
                "class": comparison["F5"].value,
                "spec": comparison["F6"].value,
                "role": comparison["F7"].value,
                "new_effects": comparison["F8"].value,
                "priority_score": comparison["F9"].value,
            },
            "verdict": comparison["A12"].value,
        },
        "cached_formula_errors": inventory["workbook"]["cached_formula_error_cells"],
        "limitations": [
            "The uploaded workbook does not contain an approved, fully populated canonical 25-player roster.",
            "This fixture freezes the observed saved state only and must not be treated as the final 25-player regression fixture.",
            "A player name is present in slot 1 without class/spec; the slot is not counted by the workbook.",
        ],
    }


def build_markdown_report(inventory: dict[str, Any], table_exports: list[dict[str, Any]], fixture: dict[str, Any]) -> str:
    workbook = inventory["workbook"]
    table_by_name = {item["table"]: item for item in table_exports}
    summary_rows = [
        ("Workbook SHA-256", inventory["source"]["sha256"]),
        ("Sheets", workbook["sheet_count"]),
        ("Excel tables", len(table_exports)),
        ("Formula cells", workbook["total_formula_count"]),
        ("Cached formula errors", workbook["total_cached_formula_errors"]),
        ("Class/spec combinations", table_by_name.get("RawCombinationsTable", {}).get("data_rows", "n/a")),
        ("Conceptual effects", table_by_name.get("EffectsReferenceTable", {}).get("data_rows", "n/a")),
        ("Saved counted players", fixture["summary"]["counted_players"]),
        ("Saved target size", fixture["context"]["target_size"]),
        ("Workbook timeline events", fixture["log_summary"]["workbook_timeline_events"]),
    ]
    lines = [
        "# Workbook v9 — frozen baseline",
        "",
        "The source workbook was read without resaving it. All exports are reproducible from the archived binary.",
        "",
        "## Summary",
        "",
        "| Metric | Value |",
        "|---|---:|",
    ]
    lines.extend(f"| {name} | {value} |" for name, value in summary_rows)
    lines.extend(
        [
            "",
            "## Known baseline risks",
            "",
            f"- `openpyxl_resave_safe`: **{workbook['openpyxl_resave_safe']}**. The workbook contains extension features that openpyxl warns it would remove on save.",
            f"- Cached Excel formula errors: **{workbook['total_cached_formula_errors']}**. They are preserved as evidence, not corrected automatically.",
            "- No approved fully populated 25-player roster is stored in the uploaded workbook; the generated fixture is the observed saved state only.",
            f"- Timeline count remains unresolved: workbook={fixture['log_summary']['workbook_timeline_events']}, project document={fixture['log_summary']['project_document_alternate_timeline_events']}, difference={fixture['log_summary']['difference']}.",
            "- Workbook v9 has no defined names; Excel/Python exchange contracts therefore still need named tables/ranges in v10.",
            "",
            "## Sheet inventory",
            "",
            "| Sheet | State | Used range | Formulas | Tables | Cached errors |",
            "|---|---|---:|---:|---|---:|",
        ]
    )
    for sheet in inventory["sheets"]:
        tables = ", ".join(item["name"] for item in sheet["tables"]) or "—"
        errors = sum(sheet["cached_formula_errors"].values())
        used = f"{sheet['max_row']} x {sheet['max_column']}"
        lines.append(
            f"| {sheet['name']} | {sheet['state']} | {used} | {sheet['formula_count']} | {tables} | {errors} |"
        )
    lines.extend(
        [
            "",
            "## Exported versioned tables",
            "",
            "| Sheet | Table | Rows | Columns | SHA-256 |",
            "|---|---|---:|---:|---|",
        ]
    )
    for item in table_exports:
        lines.append(
            f"| {item['sheet']} | {item['table']} | {item['data_rows']} | {item['columns']} | `{item['sha256']}` |"
        )
    lines.append("")
    return "\n".join(lines)


def freeze_workbook_baseline(
    workbook_path: Path,
    output_dir: Path,
    *,
    additional_sources: Iterable[Path] = (),
) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    inventory, formula_wb, value_wb = build_inventory(workbook_path)
    inventory["source"] = asdict(freeze_file(workbook_path, relative_to=output_dir.parent))
    table_exports = export_tables(formula_wb, value_wb, output_dir / "tables")
    fixture = build_default_fixture(inventory["source"]["sha256"], value_wb, inventory)

    manifest = {
        "schema_version": 1,
        "generated_at": datetime.now().astimezone().isoformat(),
        "files": [
            asdict(freeze_file(path, relative_to=output_dir.parent))
            for path in [workbook_path, *additional_sources]
            if path.exists()
        ],
        "table_exports": table_exports,
    }

    inventory_path = output_dir / "workbook_inventory.json"
    fixture_path = output_dir / "workbook_v9_observed_default_state.json"
    manifest_path = output_dir / "source_manifest.json"
    report_path = output_dir / "BASELINE_REPORT.md"
    inventory_path.write_text(json.dumps(inventory, ensure_ascii=False, indent=2), encoding="utf-8")
    fixture_path.write_text(json.dumps(fixture, ensure_ascii=False, indent=2), encoding="utf-8")
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(build_markdown_report(inventory, table_exports, fixture), encoding="utf-8")
    return {
        "inventory": inventory_path,
        "fixture": fixture_path,
        "manifest": manifest_path,
        "report": report_path,
    }
